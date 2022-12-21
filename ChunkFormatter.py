import openpyxl as op
import os
import openpyxl.styles
from openpyxl.styles import Border, Side, Alignment
import openpyxl.utils as u

# 7 Formats chunks to be ready for distribution

files = os.listdir('../MasterTracking_Data/3Catagories/Chunks/')

thin = Side(border_style="thin", color="000000")#Border style, color
border = Border(left=thin, right=thin, top=thin, bottom=thin)
#coloring
green = op.styles.PatternFill(start_color='58F139', end_color='58F139', fill_type='solid')
red = op.styles.PatternFill(start_color='F14534', end_color='F14534', fill_type='solid')
yellow = op.styles.PatternFill(start_color='F7FE28', end_color='F7FE28', fill_type='solid')
purple = op.styles.PatternFill(start_color='CB2EBF', end_color='CB2EBF', fill_type='solid')

# Generate Dictionary with Column Names
# Key values are header names
template = op.load_workbook(f'../MasterTracking_Data/3Catagories/Chunks/{files[0]}')['Sheet1']
ColNames = {}
for cell in template[1]:
    ColNames[cell.value] = cell.column_letter

finalCols = ['MOBILIZE ID FINAL', 'WORKDAY ID FINAL', 'FIRST NAME FINAL', 'LAST NAME FINAL', 'MIDDLE NAME FINAL', 'PHONE FINAL', 'EMAIL FINAL',
             'CLASS FINAL', 'SPECIALTY FINAL', 'DOB FINAL', 'GENDER FINAL', 'ADDRESS FINAL', 'CITY FINAL', 'STATE FINAL', 'ZIP FINAL',
             'LICENSE STATE FINAL', 'LICENSE NUMBER FINAL', 'SSN FINAL']
finalCols = [ColNames[x] for x in finalCols]

existCols = ['Profile Exists_AM', 'Profile Exists_BE', 'Profile Exists_WD', 'Profile Exists_EOC']
existCols = [ColNames[x] for x in existCols]

AMCols = ['EID_AM', 'WDID_AM', 'FirstName_AM', 'LastName_AM', 'Phone_AM', 'Email_AM', 'Class_AM', 'Specialty_AM', 'Gender_AM', 'City_AM', 'State_AM', 'Zip_AM', 'LicenseState_AM']
BECols = ['FirstName_BE', 'LastName_BE', 'MiddleName_BE', 'Phone_BE', 'Email_BE', 'Class_BE', 'Specialty_BE', 'DOB_BE', 'Gender_BE', 'Address_BE', 'City_BE', 'State_BE',
          'Zip_BE', 'LicenseState_BE', 'LicenseNumber_BE', 'SS#_BE']
WDCols = ['EffectiveDate_WD', 'WDID_WD', 'FirstName_WD', 'LastName_WD', 'MiddleName_WD', 'Phone_WD', 'Email_WD', 'DOB_WD', 'Address_WD', 'City_WD', 'State_WD', 'Zip_WD', 'SS#_WD']
EOCCols = ['FirstName_EOC', 'LastName_EOC', 'MiddleName_EOC', 'Phone_EOC', 'Email_EOC', 'Class_EOC', 'DOB_EOC', 'Gender_EOC', 'LicenseNumber_EOC']

AMCols = [ColNames[x] for x in AMCols]
BECols = [ColNames[x] for x in BECols]
WDCols = [ColNames[x] for x in WDCols]
EOCCols = [ColNames[x] for x in EOCCols]

existCorrespondenceTemp = {'Profile Exists_AM': AMCols, 'Profile Exists_BE': BECols, 'Profile Exists_WD': WDCols, 'Profile Exists_EOC': EOCCols}
existCorrespondence = {}

for col in existCorrespondenceTemp:
    existCorrespondence[ColNames[col]] = existCorrespondenceTemp[col]

for file in files:

    wb = op.load_workbook(f'../MasterTracking_Data/3Catagories/Chunks/{file}')
    ws = wb['Sheet1']

    # Sets dimensions
    ws.row_dimensions[1].height = 60
    for i in range(ws.max_column):
        if i < 5:
            ws.column_dimensions[u.get_column_letter(i + 1)].width = 10 # Have the profile exists be slightly smaller
        else:
            ws.column_dimensions[u.get_column_letter(i + 1)].width = 15.7 # For some reason this produces size 15 in Excel

    # Makes headers have wraptext
    for cell in ws[1]:
        cell.alignment = Alignment(wrapText=True)

    #right aligning the numbers in the IDs that are next to the email
    rightAlign_columns = [ColNames['EID_M'],ColNames['EID_AM']]
    for col in rightAlign_columns:
        for cell in ws[col]:
            cell.alignment = Alignment(horizontal='right')

    #highlighting the columns
    for col in finalCols:
        for cell in ws[col]:
            if cell.value is None:
                cell.fill = red
            else:
                cell.fill = green

    #if the 'Okay?' Column has correct, then the col before it should be green fill
    #used in the cases that all the sources are null
    for col in finalCols:
        for final_cell in ws[col]:
            okay_cell = ws.cell(row=final_cell.row, column=u.column_index_from_string(col) + 1).value
            if okay_cell == 'Correct':
                final_cell.fill = green

    #making the falses in profile exists as yellow
    for col in existCols:
        for cell in ws[col]:
            if cell.value == 'False':
                cell.fill = yellow

    #if that profile exists is false, the respective boxes for that source turn purple
    for row in ws:
        for existCol in existCols:
            existCell = row[u.column_index_from_string(existCol) - 1].value
            if existCell == 'False':
                for col in existCorrespondence[existCol]:
                    cell = row[u.column_index_from_string(col) - 1]
                    cell.fill = purple

    #adding borders to the relevant columns
    for row in ws:
        for cell in row:
            cell.border = border

    #hiding Class_EOC causing confusion
    for col in [None, 'Class_EOC', 'RowCorrect?', 'ProfileActive_M']:
        ws.column_dimensions[ColNames[col]].hidden = True

    #freezing first few columns and top row
    ws.freeze_panes = 'H2'

    print('File', file, 'Generated')
    wb.save(f'../MasterTracking_Data/3Catagories/Chunks/{file}')
