import openpyxl as op
import os
import openpyxl.styles
from openpyxl.styles import Border, Side, Alignment
import openpyxl.utils as u

# 7 part two, for Mobilize Delete Chunks

files = os.listdir('../MasterTracking_Data/MobilizeToDelete/Delete Chunks/')

thin = Side(border_style="thin", color="000000")#Border style, color
border = Border(left=thin, right=thin, top=thin, bottom=thin)
#coloring
green = op.styles.PatternFill(start_color='58F139', end_color='58F139', fill_type='solid')
red = op.styles.PatternFill(start_color='F14534', end_color='F14534', fill_type='solid')
yellow = op.styles.PatternFill(start_color='F7FE28', end_color='F7FE28', fill_type='solid')
purple = op.styles.PatternFill(start_color='CB2EBF', end_color='CB2EBF', fill_type='solid')

# Generate Dictionary with Column Names
# Key values are header names
template = op.load_workbook(f'../MasterTracking_Data/MobilizeToDelete/Delete Chunks/{files[0]}')['Sheet1']
ColNames = {}
for cell in template[1]:
    ColNames[cell.value] = cell.column_letter

for file in files:

    wb = op.load_workbook(f'../MasterTracking_Data/MobilizeToDelete/Delete Chunks/{file}')
    ws = wb['Sheet1']

    # Sets dimensions
    ws.row_dimensions[1].height = 60
    for i in range(ws.max_column):
        ws.column_dimensions[u.get_column_letter(i + 1)].width = 15.7 # For some reason this produces size 15 in Excel

    # Makes headers have wraptext
    for cell in ws[1]:
        cell.alignment = Alignment(wrapText=True)

    #right aligning the numbers in the IDs that are next to the email
    rightAlign_columns = [ColNames['Profile ID'],ColNames['WorkdayID']]
    for col in rightAlign_columns:
        for cell in ws[col]:
            cell.alignment = Alignment(horizontal='right')

    #adding borders to the relevant columns
    for row in ws:
        for cell in row:
            cell.border = border

    ws.column_dimensions[ColNames[None]].hidden = True

    print('File', file, 'Generated')
    wb.save(f'../MasterTracking_Data/MobilizeToDelete/Delete Chunks/{file}')
